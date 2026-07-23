import glob
import openpyxl
import os

from pathlib import Path


NOTEBOOK_REPO_PATH = '../$Private/Customers/$Current/DMA/Repo/Notebooks'


def process():
	notebook_repo_list = load_notebook_repo()

	notebook_delete_list = []

	path = '$DMA.xlsx'
	wb_obj = openpyxl.load_workbook(path)
	sheet_obj = wb_obj['Searches & Alerts']

	m_row = sheet_obj.max_row
	for i in range(2, m_row + 1):
		notebook_name_cell = sheet_obj.cell(row=i, column=2)
		app_cell = sheet_obj.cell(row=i, column=4)
		sharing_cell = sheet_obj.cell(row=i, column=6)
		runs_cell = sheet_obj.cell(row=i, column=9)
		notebook_name = str(notebook_name_cell.value)
		app = str(app_cell.value)
		sharing = str(sharing_cell.value)
		runs = str(runs_cell.value)

		print(f'Notebook Name: {notebook_name}')
		print(f'App:           {app}')
		print(f'Sharing:       {sharing}')
		print(f'Runs:          {runs}')

		if notebook_name not in notebook_repo_list:
			print(f'Skip notebook: not in repo!')
			continue

		if app != 'Fifth_Third_Global':
			print('Delete notebook: wrong app')
			notebook_delete_list.append(notebook_name)
			continue

		# if sharing == 'Private':
		# 	print('Delete notebook: Private')
		# 	notebook_delete_list.append(notebook_name)
		# 	continue
		#
		# if runs == "0":
		# 	print('Delete notebook: no runs')
		# 	notebook_delete_list.append(notebook_name)
		# 	continue

	print(f'Number of notebooks to be deleted: {len(notebook_delete_list)}')

	print('Notebooks that can be deleted from repo:')
	for notebook_name in notebook_delete_list:
		print(notebook_name)


def load_notebook_repo():
	results = []
	for filename in glob.glob(NOTEBOOK_REPO_PATH + '/*'):
		if os.path.isfile(filename):
			if 'metadata' not in filename:
				file_stem = Path(filename).stem
				normalized_file_stem = file_stem.replace('[ALERT] - ', '')
				normalized_file_stem = normalized_file_stem.replace('[REPORT] - ', '')
				normalized_file_stem = normalized_file_stem.replace('[SEARCH] - ', '')
				results.append(normalized_file_stem)

	# print('DEBUG: notebook repo results:')
	# for result in results:
	# 	print(result)
	# exit(9999)

	return results


if __name__ == '__main__':
	process()
