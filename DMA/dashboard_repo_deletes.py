import glob
import openpyxl
import os

from pathlib import Path


DASHBOARD_REPO_PATH = '../../$Private/Customers/$Current/DMA/Repo/Dashboards'


def process():
	dashboard_repo_list = load_dashboard_repo()

	dashboard_delete_list = []

	path = '$DMA.xlsx'
	wb_obj = openpyxl.load_workbook(path)
	# sheet_obj = wb_obj.active
	sheet_obj = wb_obj['Dashboards']

	m_row = sheet_obj.max_row
	for i in range(2, m_row + 1):
		dashboard_name_cell = sheet_obj.cell(row=i, column=2)
		app_cell = sheet_obj.cell(row=i, column=5)
		sharing_cell = sheet_obj.cell(row=i, column=7)
		views_cell = sheet_obj.cell(row=i, column=8)
		dashboard_name = str(dashboard_name_cell.value)
		app = str(app_cell.value)
		sharing = str(sharing_cell.value)
		views = str(views_cell.value)

		print(f'Dashboard Name: {dashboard_name}')
		print(f'App:            {app}')
		print(f'Sharing:        {sharing}')
		print(f'Views:          {views}')

		if dashboard_name not in dashboard_repo_list:
			print(f'Skip dashboard: not in repo!')
			continue

		if app != 'Fifth_Third_Global':
			print('Delete dashboard: wrong app')
			dashboard_delete_list.append(dashboard_name)
			continue

		if sharing == 'Private':
			print('Delete dashboard: Private')
			dashboard_delete_list.append(dashboard_name)
			continue

		if views == "0":
			print('Delete dashboard: no views')
			dashboard_delete_list.append(dashboard_name)
			continue

	print(f'Number of dashboards to be deleted: {len(dashboard_delete_list)}')

	print('Dashboards that can be deleted from repo:')
	for dashboard_name in dashboard_delete_list:
		print(dashboard_name)

def load_dashboard_repo():
	results = []
	for filename in glob.glob(DASHBOARD_REPO_PATH + '/*'):
		if os.path.isfile(filename):
			if 'metadata' not in filename:
				file_stem = Path(filename).stem
				results.append(file_stem.replace('[Splunk] ', ''))

	# print('DEBUG: dashboard repo results:')
	# for result in results:
	# 	print(result)

	return results

if __name__ == '__main__':
	process()
