import json
import openpyxl
import os
import requests

# from Reuse import dynatrace_api
# from Reuse import environment


account_id = os.getenv('DYNATRACE_AUTOMATION_ACCOUNT_ID')
client_id = os.getenv('DYNATRACE_AUTOMATION_CLIENT_ID')
client_secret = os.getenv('DYNATRACE_AUTOMATION_CLIENT_SECRET')
environment_variable_source = 'New Environment Variable Names'

print('Masked environment variables:')
print(f'account_id: {account_id[:10]}*')
print(f'client_secret: {client_secret[:5]}*{client_secret[70:]}')
print(f'client_id: {client_id[:10]}*')
print(f'environment_variable_source: {environment_variable_source}')


def get_account_management_api(api_type):
    oauth_bearer_token = get_oauth_bearer_token()
    url = f'https://api.dynatrace.com/iam/v1/accounts/{account_id}/{api_type}'
    if api_type in ['environments', 'subscriptions']:
        url = f'https://api.dynatrace.com/env/v1/accounts/{account_id}/{api_type}'
    if api_type in ['time-zones', 'regions']:
        url = f'https://api.dynatrace.com/ref/v1/{api_type}'
    if api_type == 'permissions':
        url = 'https://api.dynatrace.com/ref/v1/account/permissions'

    headers = {'accept': 'application/json', 'Authorization': 'Bearer ' + str(oauth_bearer_token)}
    r = requests.get(url, headers=headers)
    if r.status_code == 200:
        # print(r.text)
        return r
    else:
        print(f'GET Request to Account Management API {api_type} Endpoint Failed:')
        print(f'Response Status Code: {r.status_code}')
        print(f'Response Reason:      {r.reason}')
        print(f'Response Text:        {r.text}')
        print("Exiting Program")
        exit(1)


def get_oauth_bearer_token():
    headers = {'Content-type': 'application/x-www-form-urlencoded'}
    r = requests.post(f'https://sso.dynatrace.com/sso/oauth2/token?grant_type=client_credentials&client_id={client_id}&client_secret={client_secret}&scope=account-idm-read account-idm-write&resource=urn:dtaccount:{account_id}', headers=headers)
    if r.status_code == 200:
        oauth_bearer_token = r.json()["access_token"]
        return oauth_bearer_token
    else:
        print(f'POST Request to Get OAuth Bearer Token Failed:')
        print(f'Response Status Code: {r.status_code}')
        print(f'Response Reason:      {r.reason}')
        print(f'Response Text:        {r.text}')
        print("Exiting Program")
        exit(1)


def process():
    users_by_group_list = load_users_by_group_list()
    # for group_tuple in users_by_group_list:
    #     print(group_tuple[0], group_tuple[1])

    groups_dict = load_groups_dict()
    # print(groups_dict)

    # users_dict = load_users_dict()
    # # print(users_dict)
    #
    for group_tuple in users_by_group_list:
        owner_name = group_tuple[0]
        user_email = group_tuple[1]

        group_name = f'ODFL Log Viewer - {owner_name} (Management Zone Group)'
        group_uuid = groups_dict[group_name]

        test_post_user_group_membership(user_email, group_uuid)
        # post_user_group_membership(user_email, group_uuid)


# def post_user_group_membership(user_email, group_uuid):
#     print(user_email, group_uuid)
#
#     oauth_bearer_token = get_oauth_bearer_token()
#
#     headers = {'accept': 'application/json', 'Content-Type': 'application/json', 'Authorization': 'Bearer ' + oauth_bearer_token}
#     try:
#         r = requests.post(f'https://api.dynatrace.com/iam/v1/accounts/{account_id}/users/{user_email}', data=[group_uuid], headers=headers)
#         if 200 < r.status_code < 299:
#             print(f'post_user_group_membership for user {user_email} response status code: {r.status_code}')
#         else:
#             if r.status_code == 400:
#                 print(f'post_user_group_membership for user {user_email} failed with response status code {r.status_code}: duplicate')
#             else:
#                 print(f'post_user_group_membership for user {user_email} failed with response status code {r.status_code}:{r.text}')
#                 r.raise_for_status()
#     except requests.exceptions.RequestException as e:
#         print(f'post_user_group_membership for user {user_email} failed with response status code: {r.status_code}')
#         print("Please check the following HTTP response to troubleshoot the issue: ")
#         print(r.text)
#         print("Exiting Program")
#         raise SystemExit(e)


def test_post_user_group_membership(user_email, group_uuid):
    oauth_bearer_token = 'test_bearer_token'
    headers = {'accept': 'application/json', 'Content-Type': 'application/json', 'Authorization': 'Bearer ' + oauth_bearer_token}
    print(f'https://api.dynatrace.com/iam/v1/accounts/{account_id}/users/{user_email}', [group_uuid], headers)


def load_users_by_group_list():
    path = 'add_users_to_groups_from_excel.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    users_by_group_list = []

    m_row = sheet_obj.max_row
    for i in range(2, m_row + 1):
        user_email_cell = sheet_obj.cell(row=i, column=1)
        user_group_cell = sheet_obj.cell(row=i, column=2)
        user_email = user_email_cell.value
        user_group = user_group_cell.value
        if user_group:
            users_by_group_list.append((user_group, user_email))

    return sorted(users_by_group_list)


# def load_users_dict():
#     users_dict = {}
#
#     users_object = get_users()
#     users = users_object.get('items')
#     for user in users:
#         user_uid = user.get('uid')
#         user_email = user.get('email')
#         users_dict[user_email] = user_uid
#
#     return users_dict
#
#
# def get_users():
#     r = get_account_management_api('users')
#     return r.json()
#
#
def load_groups_dict():
    groups_dict = {}

    groups_object = get_groups()
    groups = groups_object.get('items')
    for group in groups:
        group_uuid = group.get('uuid')
        group_name = group.get('name')
        groups_dict[group_name] = group_uuid

    return groups_dict


def get_groups():
    r = get_account_management_api('groups')
    return r.json()


if __name__ == '__main__':
    process()
