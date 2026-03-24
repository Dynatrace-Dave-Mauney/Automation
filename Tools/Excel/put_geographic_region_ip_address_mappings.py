from Reuse import dynatrace_api
from Reuse import environment

import copy
import json
import openpyxl

friendly_function_name = 'Dynatrace Automation Reporting'
env_name_supplied = environment.get_env_name(friendly_function_name)
# For easy control from IDE
# env_name_supplied = 'Prod'
env_name, env, token = environment.get_environment_for_function(env_name_supplied, friendly_function_name)


def process():
	# get_countries()
	# get_regions_of_a_country('US')
	# get_geographic_regions_ip_address_mappings()
	# exit(1234)

	path = 'ServerIPRanges.xlsx'
	wb_obj = openpyxl.load_workbook(path)
	sheet_obj = wb_obj.active

	mappings = []

	m_row = sheet_obj.max_row
	for i in range(2, m_row + 1):
		start_ip_cell = sheet_obj.cell(row=i, column=1)
		end_ip_cell = sheet_obj.cell(row=i, column=2)
		location_cell = sheet_obj.cell(row=i, column=3)

		start_ip = str(start_ip_cell.value).strip()
		end_ip = str(end_ip_cell.value).strip()
		location = str(location_cell.value).strip()

		mapping = build_geographic_regions_ip_address_mappings(start_ip, end_ip, location)
		mappings.append(mapping)

	put_geographic_regions_ip_address_mappings(mappings)


def build_geographic_regions_ip_address_mappings(start_ip, end_ip, location):
	template = {
  "ipAddressMappingRules": [
    {
      "ipAddressMappingLocation": {
        "city": "string",
        "countryCode": "string",
        "latitude": 1,
        "longitude": 1,
        "regionCode": "string"
      },
      "ipAddressRange": {
        "address": "string",
        "addressTo": "string",
      }
    }
  ]
}


	country_code = 'US'
	state_code = None
	city = None
	latitude = 0
	longitude = 0

	if location == '1425 Madison Ave':
		state_code = 'NY'
		city = 'New York (1425)'
		latitude = 40.7128
		longitude = 74.006
	else:
		if location == 'NY6':
			state_code = 'NJ'
			city = 'Secaucus (NY6)'
			latitude = 40.7880
			longitude = 74.0545
		else:
			if location == 'MSSN  (Hicksville\Oceanside)':
				state_code = 'NY'
				city = 'Hicksville'
				latitude = 40.7731
				longitude = 73.5279
			else:
				city = location

	mapping = copy.deepcopy(template['ipAddressMappingRules'][0])

	mapping['ipAddressMappingLocation']['city'] = city
	mapping['ipAddressMappingLocation']['countryCode'] = country_code
	mapping['ipAddressMappingLocation']['latitude'] = latitude
	mapping['ipAddressMappingLocation']['longitude'] = longitude
	mapping['ipAddressMappingLocation']['regionCode'] = state_code

	mapping['ipAddressRange']['address'] = start_ip
	mapping['ipAddressRange']['addressTo'] = end_ip

	return mapping


def put_geographic_regions_ip_address_mappings(mappings):
	ip_address_mapping_rules = {'ipAddressMappingRules': mappings}
	put_ip_address_mapping_rules(ip_address_mapping_rules)


def put_ip_address_mapping_rules(ip_address_mapping_rules):
	mapping_string = json.dumps(ip_address_mapping_rules)
	print(mapping_string)

	endpoint = '/api/config/v1/geographicRegions/ipAddressMappings'

	r = dynatrace_api.put_object(f'{env}{endpoint}', token, mapping_string)

	if r.status_code == 204:
		print(f'Updated Mappings')
	else:
		if r.status_code == 400:
			print(f'Update of mappings failed due to invalid input')


def get_geographic_regions_ip_address_mappings():
	endpoint = '/api/config/v1/geographicRegions/ipAddressMappings'
	r = dynatrace_api.get_without_pagination(f'{env}{endpoint}', token)
	mappings_json = r.json()
	mappings = mappings_json.get('ipAddressMappingRules')
	for mapping in mappings:
		print(mapping)


def get_regions_of_a_country(country_code):
	endpoint = f'/api/v2/rum/regions/{country_code}'
	r = dynatrace_api.get_without_pagination(f'{env}{endpoint}', token)
	regions_json = r.json()
	regions = regions_json.get('regions')
	for region in regions:
		print(region)


def get_countries():
	endpoint = '/api/v2/rum/countries'
	r = dynatrace_api.get_without_pagination(f'{env}{endpoint}', token)
	countries_json = r.json()
	countries = countries_json.get('countries')
	for country in countries:
		print(country)


if __name__ == '__main__':
	process()
