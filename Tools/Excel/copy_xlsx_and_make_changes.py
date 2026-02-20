import openpyxl
from Reuse import report_writer


def process():
	rows = []

	unique_env_list = []
	unique_function_list = []
	unique_appname_list = []
	unique_zone_list = []

	path = '$MSHS Server Inventory.xlsx'
	wb_obj = openpyxl.load_workbook(path)
	sheet_obj = wb_obj.active

	m_row = sheet_obj.max_row
	for i in range(2, m_row + 1):
		host_cell = sheet_obj.cell(row=i, column=1)
		env_cell = sheet_obj.cell(row=i, column=2)
		app_cell = sheet_obj.cell(row=i, column=3)
		ip_address_cell = sheet_obj.cell(row=i, column=4)
		os_cell = sheet_obj.cell(row=i, column=5)
		function_cell = sheet_obj.cell(row=i, column=6)
		zone_cell = sheet_obj.cell(row=i, column=7)
		winrm_cell = sheet_obj.cell(row=i, column=8)
		tier_cell = sheet_obj.cell(row=i, column=9)
		batch_cell = sheet_obj.cell(row=i, column=10)
		deployment_date_cell = sheet_obj.cell(row=i, column=11)
		oneagent_install_status_cell = sheet_obj.cell(row=i, column=12)
		oneagent_uninstall_status_cell = sheet_obj.cell(row=i, column=13)
		notes_cell = sheet_obj.cell(row=i, column=14)

		server = str(host_cell.value)
		env = str(env_cell.value).lower()
		appname = str(app_cell.value).lower()
		function = str(function_cell.value).lower()
		zone = str(zone_cell.value).lower()
		tier = str(tier_cell.value).lower()

		ip_address = str(ip_address_cell.value)
		os = str(os_cell.value)
		winrm = str(winrm_cell.value)
		batch = str(batch_cell.value)
		deployment_date = str(deployment_date_cell.value)
		oneagent_install_status = str(oneagent_install_status_cell.value)
		oneagent_uninstall_status = str(oneagent_uninstall_status_cell.value)
		notes = str(notes_cell.value)

		env = convert_env(env)
		if env not in unique_env_list:
			unique_env_list.append(env)

		function = convert_function(function)
		if function not in unique_function_list:
			unique_function_list.append(function)

		appname = convert_appname(appname)
		if appname not in unique_appname_list:
			unique_appname_list.append(appname)

		zone = convert_zone(zone)
		if zone not in unique_zone_list:
			unique_zone_list.append(zone)

		tier = convert_tier(tier)

		print('server:                           ' + server)
		print('env:                              ' + env)
		print('appname:                          ' + appname)
		print('function:                         ' + function)
		print('zone:                             ' + zone)
		print('tier:                             ' + str(tier))

		print('ip_address                        ' + ip_address)
		print('os                                ' + os)
		print('winrm                             ' + winrm)
		print('batch                             ' + str(batch))
		print('deployment_date                   ' + str(deployment_date))
		print('oneagent_install_status           ' + str(oneagent_install_status))
		print('oneagent_uninstall_status         ' + str(oneagent_uninstall_status))
		print('notes                             ' + str(notes))

	# for appname in unique_appname_list:
	# 	print(appname)

	# for function in unique_function_list:
	# 	print(function)

	# for env in unique_env_list:
	# 	print(env)

	# for zone in unique_zone_list:
	# 	print(zone)

		host_group = f'{appname}_{function}_{env}'
		security_context = appname
		network_zone = zone

		command_line = f'--set-host-group={host_group} --set-host-property=dt.security_context={appname} --set-host-tag=primary_tags.app={appname} --set-host-tag=primary_tags.function={function} --set-host-tag=primary_tags.env={env} --set-host-tag=primary_tags.tier={tier} --set-host-tag=primary_tags.zone={zone} --set-network-zone={zone} --set-monitoring-mode=infra-only --set-system-logs-access-enabled=true --set-app-log-content-access=true'

		rows.append((server, appname, function, env, host_group, zone, tier, security_context, network_zone, ip_address, os, winrm, batch, deployment_date, oneagent_install_status, oneagent_uninstall_status, notes, command_line))

	report_headers = ('server', 'appname', 'function', 'env', 'host-group', 'zone', 'tier', 'dt.security_context', 'network-zone', 'IP Address', 'OS', 'Missing WinRM', 'Batch', 'Deployment Date', 'OneAgent installation', 'Uninstalled OneAgent', 'Notes', 'Command Line')
	report_writer.write_xlsx('MSHS Server Inventory Automation.xlsx', 'Automation', report_headers, rows, header_format=None, auto_filter=(0, len(report_headers) - 1))


def convert_env(env):
	if env == 'production' or env == 'pre-production':
		env = 'prod'

	if env == 'development':
		env = 'dev'

	if env == 'staging':
		env = 'stage'

	if env in ['dr', 'prod', 'dev', 'test', 'qa', 'stage']:
		return env
	else:
		return f'INVALID ENV: {env}'


def convert_function(function):
	if function in ['app', 'db-sql', 'db-ora', 'db', 'web']:
		return function
	else:
		return f'INVALID FUNCTION: {function}'


def convert_appname(appname):
	appname = appname.replace(' ', '-')
	appname = appname.replace('_', '-')
	appname = appname.replace(',', '-')
	appname = appname.replace('(', '-')
	appname = appname.replace(')', '-')
	appname = appname.replace('--', '-')
	appname = appname.replace('--', '-')
	appname = appname.replace('/', '-')
	appname = appname.replace('\\', '-')
	appname = appname.replace('.', '-')

	if appname.startswith('-'):
		appname = appname[1:len(appname)]

	if appname.endswith('-'):
		appname = appname[0:len(appname) - 1]

	return appname


def convert_zone(zone):
	if zone in ['onprem', 'azure']:
		return zone
	else:
		return f'INVALID TIER: {zone}'


def convert_tier(tier):
	if tier == 'yes':
		return 1
	else:
		return 0


if __name__ == '__main__':
	process()
