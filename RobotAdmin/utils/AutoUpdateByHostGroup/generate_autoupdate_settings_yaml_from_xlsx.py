import openpyxl
import yaml

autoupdate_comments = '''# Update the OneAgent "auto update" setting at host or host group level
# Supported settings:
# DISABLED|ENABLED|INHERITED
# Maintenance windows are not supported
'''

def process():
	disabled_setting_host_list = []
	disabled_setting_host_group_list = []
	enabled_setting_host_list = []
	enabled_setting_host_group_list = []
	inherited_setting_host_list = []
	inherited_setting_host_group_list = []

	path = 'autoupdate.xlsx'
	wb_obj = openpyxl.load_workbook(path)
	sheet_obj = wb_obj.active

	m_row = sheet_obj.max_row
	for i in range(2, m_row + 1):
		host_group_cell = sheet_obj.cell(row=i, column=1)
		host_cell = sheet_obj.cell(row=i, column=2)
		scope_cell = sheet_obj.cell(row=i, column=6)
		setting_cell = sheet_obj.cell(row=i, column=7)
		if scope_cell.value or setting_cell.value:
			if scope_cell.value not in ['H', 'HG']:
				print('Aborting.  Row ' + str(i) + ': Invalid "Scope": ' + str(scope_cell.value))
				exit(1)
			if setting_cell.value not in ['INHERITED', 'ENABLED', 'DISABLED']:
				print('Aborting.  Row ' + str(i) + ': Invalid setting: ' + str(setting_cell.value))
				exit(2)
			if scope_cell.value == 'HG' and host_group_cell.value == 'None':
				print('Aborting.  Row ' + str(i) + ': Invalid Host Group for "HG" scope: ' + str(host_group_cell.value))
				exit(3)
			# print('scope:         ' + scope_cell.value)
			# print('setting:       ' + setting_cell.value)
			# print('host id:       ' + host_cell.value)
			# print('host group id: ' + host_group_cell.value)

			if setting_cell.value == 'DISABLED':
				if scope_cell.value == 'H':
					disabled_setting_host_list.append(host_cell.value)
				else:
					disabled_setting_host_group_list.append(host_group_cell.value)

			if setting_cell.value == 'ENABLED':
				if scope_cell.value == 'H':
					enabled_setting_host_list.append(host_cell.value)
				else:
					enabled_setting_host_group_list.append(host_group_cell.value)

			if setting_cell.value == 'INHERITED':
				if scope_cell.value == 'H':
					inherited_setting_host_list.append(host_cell.value)
				else:
					inherited_setting_host_group_list.append(host_group_cell.value)

	# disabled_setting_dict = {'setting': 'DISABLED', 'hostgroups':  disabled_setting_host_group_list, 'hosts': disabled_setting_host_list}
	# enabled_setting_dict = {'setting': 'ENABLED', 'hostgroups':  enabled_setting_host_group_list, 'hosts': enabled_setting_host_list}
	# inherited_setting_dict = {'setting': 'INHERITED', 'hostgroups':  inherited_setting_host_group_list, 'hosts': inherited_setting_host_list}
	combined_setting_dict = {
		'settings': [
			{'setting': 'DISABLED', 'hostgroups':  disabled_setting_host_group_list, 'hosts': disabled_setting_host_list},
			{'setting': 'ENABLED', 'hostgroups':  enabled_setting_host_group_list, 'hosts': enabled_setting_host_list},
			{'setting': 'INHERITED', 'hostgroups':  inherited_setting_host_group_list, 'hosts': inherited_setting_host_list},
		]
	}

	# output_file_name = 'generated_autoupdate_disabled.yaml'
	# with open(output_file_name, 'w', encoding='utf-8') as file:
	# 	file.write(autoupdate_comments)
	# with open(output_file_name, 'a') as file:
	# 	yaml.dump(disabled_setting_dict, file, sort_keys=False)
	#
	# output_file_name = 'generated_autoupdate_enabled.yaml'
	# with open(output_file_name, 'w', encoding='utf-8') as file:
	# 	file.write(autoupdate_comments)
	# with open(output_file_name, 'a') as file:
	# 	yaml.dump(enabled_setting_dict, file, sort_keys=False)
	#
	# output_file_name = 'generated_autoupdate_inherited.yaml'
	# with open(output_file_name, 'w', encoding='utf-8') as file:
	# 	file.write(autoupdate_comments)
	# with open(output_file_name, 'a') as file:
	# 	yaml.dump(inherited_setting_dict, file, sort_keys=False)
	#
	output_file_name = 'autoupdate.yaml'
	with open(output_file_name, 'w', encoding='utf-8') as file:
		file.write(autoupdate_comments)
	with open(output_file_name, 'a') as file:
		yaml.dump(combined_setting_dict, file, sort_keys=False)

if __name__ == '__main__':
    process()
